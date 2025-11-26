from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from .. import models
from ..database import get_db
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/api/export", tags=["export"])

@router.get("/pdf/{project_id}")
def export_pdf(project_id: int, db: Session = Depends(get_db)):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    elements = db.query(models.Element).filter(models.Element.project_id == project_id).all()
    connections = db.query(models.Connection).filter(models.Connection.project_id == project_id).all()
    panel_elements = db.query(models.PanelElement).filter(models.PanelElement.project_id == project_id).all()
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Заголовок
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, f"Проект: {project.name}")
    
    # Таблица элементов
    y_pos = height - 100
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y_pos, "Элементы схемы:")
    y_pos -= 20
    
    c.setFont("Helvetica", 10)
    headers = ["ID", "Тип", "Название", "X", "Y"]
    col_widths = [60, 80, 150, 50, 50]
    x_start = 50
    
    # Заголовки таблицы
    x = x_start
    for i, header in enumerate(headers):
        c.drawString(x, y_pos, header)
        x += col_widths[i]
    y_pos -= 15
    
    # Данные
    for element in elements:
        if y_pos < 100:
            c.showPage()
            y_pos = height - 50
        c.drawString(x_start, y_pos, element.element_id)
        c.drawString(x_start + col_widths[0], y_pos, element.type)
        c.drawString(x_start + col_widths[0] + col_widths[1], y_pos, element.name)
        c.drawString(x_start + col_widths[0] + col_widths[1] + col_widths[2], y_pos, f"{element.x:.2f}")
        c.drawString(x_start + col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3], y_pos, f"{element.y:.2f}")
        y_pos -= 15
    
    # Таблица связей
    c.showPage()
    y_pos = height - 50
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y_pos, "Связи (кабели):")
    y_pos -= 20
    
    c.setFont("Helvetica", 10)
    headers = ["От", "К", "Сечение (мм²)", "Жил", "Длина (м)"]
    col_widths = [60, 60, 80, 50, 80]
    
    x = x_start
    for i, header in enumerate(headers):
        c.drawString(x, y_pos, header)
        x += col_widths[i]
    y_pos -= 15
    
    for connection in connections:
        if y_pos < 100:
            c.showPage()
            y_pos = height - 50
        from_elem = db.query(models.Element).filter(models.Element.id == connection.from_element_id).first()
        to_elem = db.query(models.Element).filter(models.Element.id == connection.to_element_id).first()
        c.drawString(x_start, y_pos, from_elem.element_id if from_elem else str(connection.from_element_id))
        c.drawString(x_start + col_widths[0], y_pos, to_elem.element_id if to_elem else str(connection.to_element_id))
        c.drawString(x_start + col_widths[0] + col_widths[1], y_pos, f"{connection.cable_section:.2f}")
        c.drawString(x_start + col_widths[0] + col_widths[1] + col_widths[2], y_pos, str(connection.wire_count))
        c.drawString(x_start + col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3], y_pos, f"{connection.length:.2f}" if connection.length else "-")
        y_pos -= 15
    
    c.save()
    buffer.seek(0)
    
    return Response(content=buffer.read(), media_type="application/pdf", 
                   headers={"Content-Disposition": f"attachment; filename=project_{project_id}.pdf"})

@router.get("/excel/{project_id}")
def export_excel(project_id: int, db: Session = Depends(get_db)):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    elements = db.query(models.Element).filter(models.Element.project_id == project_id).all()
    connections = db.query(models.Connection).filter(models.Connection.project_id == project_id).all()
    panel_elements = db.query(models.PanelElement).filter(models.PanelElement.project_id == project_id).all()
    
    wb = openpyxl.Workbook()
    
    # Лист элементов
    ws_elements = wb.active
    ws_elements.title = "Элементы"
    headers = ["ID", "Тип", "Название", "X", "Y", "Свойства"]
    ws_elements.append(headers)
    
    # Стиль заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_elements[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for element in elements:
        ws_elements.append([
            element.element_id,
            element.type,
            element.name,
            element.x,
            element.y,
            str(element.properties)
        ])
    
    # Лист связей
    ws_connections = wb.create_sheet("Связи")
    headers = ["От элемента", "К элементу", "Сечение (мм²)", "Количество жил", "Длина (м)"]
    ws_connections.append(headers)
    
    for cell in ws_connections[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for connection in connections:
        from_elem = db.query(models.Element).filter(models.Element.id == connection.from_element_id).first()
        to_elem = db.query(models.Element).filter(models.Element.id == connection.to_element_id).first()
        ws_connections.append([
            from_elem.element_id if from_elem else str(connection.from_element_id),
            to_elem.element_id if to_elem else str(connection.to_element_id),
            connection.cable_section,
            connection.wire_count,
            connection.length if connection.length else ""
        ])
    
    # Лист элементов щита
    ws_panel = wb.create_sheet("Элементы щита")
    headers = ["ID элемента", "Позиция X", "Позиция Y", "Ширина", "Высота"]
    ws_panel.append(headers)
    
    for cell in ws_panel[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for panel_elem in panel_elements:
        elem = db.query(models.Element).filter(models.Element.id == panel_elem.element_id).first()
        ws_panel.append([
            elem.element_id if elem else str(panel_elem.element_id),
            panel_elem.position_x,
            panel_elem.position_y,
            panel_elem.width,
            panel_elem.height
        ])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(content=buffer.read(), 
                   media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": f"attachment; filename=project_{project_id}.xlsx"})

